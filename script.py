# -*- coding: utf-8 -*-
"THIS IS A XML CONVERTER to Invoice"
import openpyxl
from datetime import datetime
from xml.etree.ElementTree import Element, SubElement, tostring, register_namespace, QName
from xml.dom import minidom
from openpyxl import load_workbook

def create_ubl_xml(data):
    # Create UBL XML document
    common_namespace = "urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2"
    basic_namespace = "urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2"
    
    ubl_doc = Element(f'{{{basic_namespace}}}CreditNote')
    # Register the 'cbc' and 'cac' namespace prefix
    register_namespace('cbc', basic_namespace) #cbc 
    register_namespace('cac', common_namespace) #cac
    #incremental id variable
    invoice_id = 1
    # Create elements and add data from Excel
    for row in data:
        invoice_line = SubElement(ubl_doc, f'{{{common_namespace}}}CreditNoteLine')
        invoice_id_element = SubElement(invoice_line, QName(basic_namespace, 'ID'))
        invoice_id_element.text = str(invoice_id)
        #increase the id 
        invoice_id += 1
        #note
        note_element = SubElement(invoice_line, 'cbc:Note')
        note_element.text = str(row['Note'])
        # CreditedQuantity
        invoiced_quantity = SubElement(invoice_line, f'{{{basic_namespace}}}CreditedQuantity', unitCode="KWT")
        invoiced_quantity.text = str(row['CreditedQuantity']) + ".00"
        # LineExtensionAmount
        line_extension_amount = SubElement(invoice_line, f'{{{basic_namespace}}}LineExtensionAmount', currencyID="EUR")
        lineExtension = str(row['LineExtensionAmount'])
        line_extension_amount.text = lineExtension
        # InvoicePeriod
        additional_item_InvoicePeriod_element = SubElement(invoice_line, QName(common_namespace, 'InvoicePeriod'))
        item_InvoicePeriod = SubElement(additional_item_InvoicePeriod_element, f'{{{basic_namespace}}}StartDate')
        item_InvoicePeriod.text = str(row['StartDate'])
        item_InvoicePeriod = SubElement(additional_item_InvoicePeriod_element, f'{{{basic_namespace}}}EndDate')
        item_InvoicePeriod.text = str(row['EndDate'])
        #Standart Values 
        item_element = SubElement(invoice_line, QName(common_namespace, 'Item'))
        name_element = SubElement(item_element, QName(basic_namespace, 'Name'))
        name_element.text = "Termo Energia Variável (1)"
        classified_tax_category_element = SubElement(item_element, QName(common_namespace, 'ClassifiedTaxCategory')) 
        #id
        id_element = SubElement(classified_tax_category_element, QName(basic_namespace, 'ID'))
        id_element.text = "S"
        #percentagem
        percent_element = SubElement(classified_tax_category_element, QName(basic_namespace, 'Percent'))
        percent_element.text = "23.00"
        tax_scheme_element = SubElement(classified_tax_category_element, QName(common_namespace, 'TaxScheme'))
        id_tax_scheme_element = SubElement(tax_scheme_element, QName(basic_namespace, 'ID'))
        id_tax_scheme_element.text = "VAT"
        #Additional XML structure
        additional_properties = [
            ("CICLO_HORARIO", "SEMANAL"), #SEM CICLO 
            ("TARIFARIO", "TETRA_HORARIO"), #SIMPLES
            ("TENSAO",  str(row['Tensao'])),
            ("TIPO_LEITURA", "REAL"),
            ("FATOR", "1"),
            ("CPE", str(row['Value']))
        ]
        #cycle to additional properties
        for name, value in additional_properties:
            additional_item_property_element = SubElement(item_element, QName(common_namespace, 'AdditionalItemProperty'))
            name_property_element = SubElement(additional_item_property_element, QName(basic_namespace, 'Name'))
            name_property_element.text = name
            value_element = SubElement(additional_item_property_element, QName(basic_namespace, 'Value'))
            value_element.text = value
        #PriceAmount
        additional_item_price_element = SubElement(invoice_line, QName(common_namespace, 'Price'))
        item_PriceAmount = SubElement(additional_item_price_element, f'{{{basic_namespace}}}PriceAmount', currencyID="EUR")
        item_PriceAmount.text = str(row['PriceAmount'])
    return ubl_doc
#save xml to file (.xml)
def save_xml_to_file(xml_data, filename):
    # Save the XML to a file
    with open(filename, 'w') as xml_file:
        xml_file.write(prettify(xml_data))
# Prettify the XML output
def prettify(elem):
    rough_string = tostring(elem, 'utf-8').decode('utf-8')
    reparsed = minidom.parseString(rough_string)
    return reparsed.toprettyxml(indent="  ")
#convert excel to xml 
def excel_to_ubl(input_excel, output_xml):
    # Load Excel data
    workbook = load_workbook(input_excel, data_only=True)
    sheet = workbook.active
    # Extract data from Excel
    excel_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
         # Format the date to "yyyy-mm-dd"
        dateStart_value = "2022-01-01"
        formatted_date = datetime.strptime(dateStart_value, "%Y-%m-%d").strftime("%Y-%m-%d")
        dateEnd_value = "2024-12-31"
        formatted_date1 = datetime.strptime(dateEnd_value, "%Y-%m-%d").strftime("%Y-%m-%d")
        #Columns mapping 
        price_amount = str(row[7])
        excel_data.append({
            'Note': row [2],
            'Tensao': row [1],
            'CreditedQuantity': row [4],
            'LineExtensionAmount': row[6],
            'Value': row[2],
            'PriceAmount': price_amount,
            'StartDate': formatted_date,
            'EndDate': formatted_date1
        })
    # Create UBL XML document
    ubl_xml = create_ubl_xml(excel_data)
    # Save UBL XML to file
    save_xml_to_file(ubl_xml, output_xml)
if __name__ == "__main__":
    # Specify the input Excel file and output UBL XML file
    input_excel_file = r"C:\Users\AG06878\Desktop\Teste_Excel_Xml\faturas_new.xlsx"
    output_ubl_xml_file = r"C:\Users\AG06878\Desktop\Teste_Excel_Xml\output.xml"
    # Convert Excel to UBL XML
    excel_to_ubl(input_excel_file, output_ubl_xml_file)
